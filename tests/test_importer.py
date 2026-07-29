from datetime import date

import pytest
from openpyxl import Workbook

from shopfloor_importer.duplicates import SubmissionLedger, fingerprint
from shopfloor_importer.config import load_config
from shopfloor_importer.model import FieldSpec
from shopfloor_importer.submit import fill_enabled_field, submit_browser
from shopfloor_importer.workbook import read_workbook
from examples.generate_workbooks import save_measurements, save_work_orders


FIELDS = (
    FieldSpec("Ref", "externalReference", True),
    FieldSpec("Qty", "plannedQuantity", True, "integer"),
    FieldSpec("Due", "dueDate", True, "date"),
    FieldSpec("Priority", "priorityCode", choices=("LOW", "HIGH")),
)


def test_generated_examples_are_parseable_and_anonymized():
    work_orders = save_work_orders()
    measurements = save_measurements()
    assert work_orders.suffix == measurements.suffix == ".xlsx"
    assert "DEMO" in next(workbook_values(work_orders))[0]


def workbook_values(path):
    from openpyxl import load_workbook

    rows = load_workbook(path, read_only=True, data_only=True).active.iter_rows(values_only=True)
    next(rows)
    return rows


def workbook(tmp_path, headings, rows):
    book = Workbook()
    sheet = book.active
    sheet.append(headings)
    for row in rows:
        sheet.append(row)
    path = tmp_path / "input.xlsx"
    book.save(path)
    return path


def test_parses_converts_and_maps_with_stable_targets(tmp_path):
    path = workbook(tmp_path, ["Ref", "Qty", "Due", "Priority"], [[" A-1 ", 4, date(2030, 1, 2), "HIGH"]])
    record = read_workbook(path, FIELDS)[0]
    assert record.valid
    assert record.values == {"Ref": "A-1", "Qty": 4, "Due": "2030-01-02", "Priority": "HIGH"}
    assert record.mapped == {"externalReference": "A-1", "plannedQuantity": 4, "dueDate": "2030-01-02", "priorityCode": "HIGH"}


def test_missing_required_header_is_file_error(tmp_path):
    path = workbook(tmp_path, ["Ref", "Qty"], [["A", 1]])
    with pytest.raises(ValueError, match="missing required columns: Due"):
        read_workbook(path, FIELDS)


def test_reports_each_invalid_row_and_value(tmp_path):
    path = workbook(tmp_path, ["Ref", "Qty", "Due", "Priority"], [[None, 1.5, "not-date", "MEDIUM"]])
    record = read_workbook(path, FIELDS)[0]
    assert record.row_number == 2
    assert record.errors == [
        "Ref: value is required", "Qty: must be a whole number",
        "Due: Invalid isoformat string: 'not-date'", "Priority: must be one of: LOW, HIGH",
    ]


def test_reads_report_with_headers_after_preamble(tmp_path):
    book = Workbook()
    sheet = book.active
    for _ in range(19):
        sheet.append(["report metadata"])
    sheet.append(["Cav #", "T @ 90", "T @ 10"])
    sheet.append([1, 2.441, 2.436])
    path = tmp_path / "report.xlsx"
    book.save(path)
    fields = (
        FieldSpec("Cav #", "__row__", True, "integer"),
        FieldSpec("T @ 90", "#row-{row}-t90", True, "number"),
        FieldSpec("T @ 10", "#row-{row}-tpl", True, "number"),
    )
    record = read_workbook(path, fields, header_row=20)[0]
    assert record.row_number == 21
    assert record.mapped == {
        "__row__": 1,
        "#row-{row}-t90": 2.441,
        "#row-{row}-tpl": 2.436,
    }


def test_measurement_config_limits_cavity_rows_and_maps_all_grid_fields():
    config = load_config("config/measurement-browser.yaml")
    assert config.integration == "browser"
    assert config.website_url == "https://cpna_sfol.berryplastics.com"
    assert config.auth == "pending"
    assert config.header_row == 20
    assert config.row_key_column == "Cav #"
    assert config.fields[0].choices == tuple(str(value) for value in range(1, 9))
    assert [field.column for field in config.fields[1:]] == [
        "T @ 90", "T @ 10", "T avg", "T ovality", "E @ 90",
        "E @ 10", "E avg", "E ovality", "Weight",
    ]


def test_pending_authentication_blocks_submission_before_browser_launch():
    config = load_config("config/measurement-browser.yaml")
    with pytest.raises(RuntimeError, match="authentication is not configured"):
        submit_browser(config, {"__cavity_row__": 1})


def test_fingerprint_is_order_independent_and_ledger_persists(tmp_path):
    first = {"a": 1, "b": "two"}
    second = {"b": "two", "a": 1}
    assert fingerprint(first) == fingerprint(second)
    path = tmp_path / "ledger.json"
    ledger = SubmissionLedger(path)
    assert not ledger.contains(first)
    ledger.add(first)
    assert SubmissionLedger(path).contains(second)


class FakeLocator:
    def __init__(self, *, checked=False, enabled=True):
        self.checked = checked
        self.enabled = enabled
        self.filled = None
        self.waited = None

    def is_checked(self):
        return self.checked

    def check(self):
        self.checked = True

    def wait_for(self, *, state):
        self.waited = state

    def is_enabled(self):
        return self.enabled

    def fill(self, value):
        self.filled = value


class FakePage:
    def __init__(self, locators):
        self.locators = locators

    def locator(self, selector):
        return self.locators[selector]


def test_checks_enable_box_before_filling_field():
    toggle = FakeLocator()
    field = FakeLocator()
    page = FakePage({"#enable-t90": toggle, "#t90": field})
    fill_enabled_field(page, "#t90", 90, "#enable-t90")
    assert toggle.checked
    assert field.waited == "visible"
    assert field.filled == "90"


def test_refuses_to_fill_a_field_that_remains_disabled():
    page = FakePage({"#weight": FakeLocator(enabled=False)})
    with pytest.raises(RuntimeError, match="disabled after enabling"):
        fill_enabled_field(page, "#weight", 12.5)
