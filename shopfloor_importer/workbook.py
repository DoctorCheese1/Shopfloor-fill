from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .model import FieldSpec, RowRecord


def _convert(value: Any, spec: FieldSpec) -> Any:
    if spec.kind == "string":
        return str(value).strip()
    if spec.kind == "integer":
        if isinstance(value, bool) or int(value) != float(value):
            raise ValueError("must be a whole number")
        return int(value)
    if spec.kind == "number":
        return float(value)
    if spec.kind == "date":
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date):
            return value.isoformat()
        return date.fromisoformat(str(value).strip()).isoformat()
    raise ValueError(f"unsupported type {spec.kind!r}")


def read_workbook(
    path: str | Path, fields: tuple[FieldSpec, ...], header_row: int = 1
) -> list[RowRecord]:
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("input must be an .xlsx workbook")
    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    if header_row < 1:
        raise ValueError("header_row must be at least 1")
    try:
        for _ in range(header_row - 1):
            next(rows)
        headings = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        raise ValueError("workbook is empty") from None
    duplicates = sorted({h for h in headings if h and headings.count(h) > 1})
    if duplicates:
        raise ValueError(f"duplicate column headers: {', '.join(duplicates)}")
    missing = [spec.column for spec in fields if spec.required and spec.column not in headings]
    if missing:
        raise ValueError(f"missing required columns: {', '.join(missing)}")
    positions = {name: index for index, name in enumerate(headings)}
    output: list[RowRecord] = []
    for number, row in enumerate(rows, header_row + 1):
        if all(value is None or str(value).strip() == "" for value in row):
            continue
        record = RowRecord(number, {})
        for spec in fields:
            raw = row[positions[spec.column]] if spec.column in positions and positions[spec.column] < len(row) else None
            if raw is None or str(raw).strip() == "":
                if spec.required:
                    record.errors.append(f"{spec.column}: value is required")
                continue
            try:
                value = _convert(raw, spec)
                if spec.choices and str(value) not in spec.choices:
                    raise ValueError(f"must be one of: {', '.join(spec.choices)}")
                record.values[spec.column] = value
                record.mapped[spec.target] = value
            except (TypeError, ValueError) as error:
                record.errors.append(f"{spec.column}: {error}")
        output.append(record)
    return output
